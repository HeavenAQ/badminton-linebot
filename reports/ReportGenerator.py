from datetime import datetime as dt
from typing import Any
import io
from firebase_admin import credentials, json
from firebase_admin import firestore
import firebase_admin
from google.cloud.firestore import DocumentSnapshot
from google.cloud.firestore_v1.stream_generator import StreamGenerator
from openpyxl import Workbook
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
import pandas as pd

from Types import PortfolioRecord, Student


class FirestoreHandler:
    def __init__(self, firebase_cred_path: bytes, root_collection: str):
        self.__firebase_cred_path = firebase_cred_path
        self.__root_collection = root_collection
        self.docs = self.init_docs()

    def init_docs(self) -> StreamGenerator[DocumentSnapshot]:
        cred_json = json.loads(self.__firebase_cred_path.decode("utf-8"))
        cred = credentials.Certificate(cred_json)
        app = firebase_admin.initialize_app(cred)
        db = firestore.client(app)
        return db.collection(self.__root_collection).stream()


class ReportGenerator:
    def __init__(
        self, firestore_credentials: bytes, root_collection: str, output_path: str
    ):
        # for firebase admin sdk
        self.firestore = FirestoreHandler(firestore_credentials, root_collection)

        # for workbook and openpyxl
        self.workbook = WorkbookHandler(output_path)

        # for student data
        self.students = [*map(self.extract_student_data, self.firestore.docs)]

    def init_workbook(self):
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        return wb

    def extract_student_data(self, doc: DocumentSnapshot) -> Student:
        doc_dict = doc.to_dict()
        if not doc_dict:
            raise ValueError("Document is empty")

        date_format = "%Y-%m-%d-%H-%M"
        clear_start_date = dt.strptime("2024-11-04-00-00", date_format)
        student_data = Student(
            {
                "name": doc_dict["Name"],
                "line_id": doc_dict["Id"],
                "handedness": "right" if doc_dict["Handedness"] == 1 else "left",
                "portfolio": [
                    PortfolioRecord(
                        {
                            "date": record["DateTime"],
                            "skill": "serve"
                            if clear_start_date
                            > dt.strptime(record["DateTime"], date_format)
                            else "clear",
                            "score": record["Rating"],
                            "ai_note": record["AINote"],
                            "preview_note": ""
                            if "尚未填寫" in (note := record["PreviewNote"])
                            else note,
                            "reflection": ""
                            if "尚未填寫" in (note := record["Reflection"])
                            else note,
                        }
                    )
                    for record in (
                        doc_dict["Portfolio"]["Serve"] | doc_dict["Portfolio"]["Clear"]
                    ).values()
                ],
            }
        )
        # sort by date to ensure the order is correct
        student_data["portfolio"].sort(
            key=lambda x: dt.strptime(
                x["date"],
                "%Y-%m-%d-%H-%M",
            )
        )
        return student_data

    def get_studemtn_missing_fields(self, sheet: Student) -> list[str]:
        # initialize the list
        student_missing_fields = []
        action_dict = {
            "reflection": "「本週學習反思」",
            "preview_note": "「課前動作檢測」",
        }

        # show the students who haven't completed the reflection and/or preview_note
        for record in sheet["portfolio"]:
            missing_fields = []
            if not record["reflection"]:
                missing_fields.append("reflection")
            if not record["preview_note"]:
                missing_fields.append("preview_note")

            if missing_fields:
                student_missing_fields.append(
                    f"{sheet['name']}：「{record['date']}」- {'及'.join(map(lambda x: action_dict[x], missing_fields))}"
                )
        return student_missing_fields

    def create_student_progress_chart(
        self, portfolio_df: pd.DataFrame, skills: list[str]
    ) -> io.BytesIO:
        # Create a figure with n subplots
        _, axs = plt.subplots(len(skills), 1, figsize=(6, 8))
        for i, skill in enumerate(skills):
            # Plot the skills
            current_df = portfolio_df[portfolio_df["skill"] == skill]
            axs[i].plot(
                current_df["date"].apply(lambda x: "/".join(x.split("-")[1:3])),
                current_df["score"],
                marker="o",
                color="b",
                label=f"{skill.capitalize()} Score",
            )

            # Set the title, x-axis label, and y-axis label
            axs[i].set_title(f"{skill.capitalize()} Skill Progress")
            axs[i].set_xlabel("Date")
            axs[i].set_ylabel("Score")

        plt.tight_layout()  # for better spacing
        img_stream = io.BytesIO()
        plt.savefig(img_stream, format="png")
        plt.close()
        img_stream.seek(0)
        return img_stream

    def generate_students_records(self):
        for student in self.students:
            # skip the admin and the teacher
            if student["name"] == "Heaven" or "林國欽" in student["name"]:
                continue

            # show the students who haven't completed the reflection and/or preview_note
            missing_fields = self.get_studemtn_missing_fields(student)
            if missing_fields:
                print(missing_fields)
                # send_line_warning(sheet["line_id"], sheet["name"], missing_fields

            # Convert the student portfolio records to a DataFrame
            portfolio_df = pd.DataFrame(student["portfolio"])

            # Create a new sheet for the student
            ws = self.workbook.create_student_sheet(student)
            self.workbook.add_student_records(ws, portfolio_df)
            self.workbook.mark_missing_fields(ws)

            # create a progress chart for the student
            img_stream = self.create_student_progress_chart(
                portfolio_df, ["serve", "clear"]
            )
            img = Image(img_stream)
            ws.add_image(img, "A20")

        self.workbook.save()

    def generate_average_and_median_score_report(self, specified_dates: list[str]):
        # List of specified dates in mm/dd format
        # Convert specified_dates to a set for faster lookup
        specified_dates_set = set(specified_dates)

        # Define the date cutoff as 11/04
        date_cutoff = dt.strptime("11/04", "%m/%d")

        # Initialize an empty list to collect all records
        all_records = []

        # Collect all student names
        all_student_names = set()

        # Dictionary to keep track of dates each student has records for
        student_dates = {}

        for student in self.students:
            # Skip the admin and the teacher
            if student["name"] == "Heaven" or "林國欽" in student["name"]:
                continue

            all_student_names.add(student["name"])
            student_dates.setdefault(student["name"], set())

            # For each portfolio record, extract date, score, and skill
            for name in student["portfolio"]:
                date_str = name["date"]  # Format is "%Y-%m-%d-%H-%M"
                skill = name["skill"]
                # Parse date string
                try:
                    date_obj = dt.strptime(date_str, "%Y-%m-%d-%H-%M")
                except ValueError:
                    # Handle any parsing errors
                    continue

                # Filter to include only records on or after 11/04 for "clear" skill
                if skill == "clear" and date_obj < date_cutoff:
                    continue

                # Extract month and day
                month_day = date_obj.strftime("%m/%d")
                if month_day in specified_dates_set:
                    all_records.append(
                        {
                            "date": month_day,
                            "score": name["score"],
                            "skill": skill,
                            "name": student["name"],
                        }
                    )
                    # Record that this student has a record on this date
                    student_dates[student["name"]].add(month_day)

        # Convert to pandas DataFrame
        records_df = pd.DataFrame(all_records)

        if records_df.empty:
            print("No records found for the specified dates.")
            return

        # Identify students who are missing any records on the specified dates
        students_missing_dates = set()
        for student_name in all_student_names:
            dates_with_records = student_dates.get(student_name, set())
            missing_dates = specified_dates_set - dates_with_records
            if missing_dates:
                students_missing_dates.add(student_name)
        print(f"Students missing records: {students_missing_dates}")

        # Create a sheet in the workbook
        ws = self.workbook.wb.create_sheet("Average and Median Scores")

        # Create a sheet to save the list of students who did not have records on the dates
        missing_ws = self.workbook.wb.create_sheet("Students Missing Records")
        missing_ws.append(["Student Name", "Missing Dates"])
        missing_ws_df = {"Student Name": [], "Missing Dates": []}

        # Write the students and their missing dates to the worksheet
        for student_name in students_missing_dates:
            dates_with_records = student_dates.get(student_name, set())
            missing_dates = specified_dates_set - dates_with_records
            missing_dates_str = ", ".join(sorted(missing_dates))
            missing_ws.append([student_name, missing_dates_str])
            missing_ws_df["Student Name"].append(student_name)
            missing_ws_df["Missing Dates"].append(missing_dates_str)
        missing_ws_df = pd.DataFrame(missing_ws_df)

        print(missing_ws_df)
        missing_ws_df["Missing Dates"] = missing_ws_df["Missing Dates"].apply(
            lambda x: x.split(", ")
        )

        filtered_records_df = records_df
        for name in students_missing_dates:
            missing_dates = missing_ws_df[missing_ws_df["Student Name"] == name][
                "Missing Dates"
            ].values[0]

            for date in missing_dates:
                print(f"Removing records for {name} on {date}")
                filtered_records_df = filtered_records_df[
                    ~(
                        (filtered_records_df["name"] == name)
                        & (filtered_records_df["date"] == date)
                    )
                ]

        # Process each skill separately
        for skill in ["serve", "clear"]:
            skill_df = filtered_records_df[filtered_records_df["skill"] == skill]

            if skill_df.empty:
                continue

            # **Keep only the highest score per student per date**
            skill_df = skill_df.groupby(["name", "date"], as_index=False)["score"].max()

            # Group by date and compute average and median scores
            avg_scores = skill_df.groupby("date")["score"].mean().reset_index()
            median_scores = skill_df.groupby("date")["score"].median().reset_index()

            # Ensure dates are sorted according to specified_dates
            avg_scores["date"] = pd.Categorical(
                avg_scores["date"], categories=specified_dates, ordered=True
            )
            avg_scores = avg_scores.sort_values("date")

            median_scores["date"] = pd.Categorical(
                median_scores["date"], categories=specified_dates, ordered=True
            )
            median_scores = median_scores.sort_values("date")

            # Write headers for the skill
            ws.append([f"{skill.capitalize()} Skill"])
            ws.append(["Date", "Average Score", "Median Score"])

            for avg_row, median_row in zip(
                avg_scores.itertuples(index=False),
                median_scores.itertuples(index=False),
            ):
                ws.append([avg_row.date, avg_row.score, median_row.score])

            # Add an empty row for spacing
            ws.append([])

            # Create a line plot of the average and median scores over time
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.plot(
                avg_scores["date"],
                avg_scores["score"],
                marker="o",
                linestyle="-",
                label="Average Score",
            )
            ax.plot(
                median_scores["date"],
                median_scores["score"],
                marker="x",
                linestyle="--",
                label="Median Score",
            )
            ax.set_title(f"{skill.capitalize()} Skill Scores Over Time")
            ax.set_xlabel("Date")
            ax.set_ylabel("Score")
            ax.set_xticks(range(len(avg_scores["date"])))
            ax.set_xticklabels(avg_scores["date"], rotation=45)
            ax.legend()

            plt.tight_layout()
            img_stream = io.BytesIO()
            plt.savefig(img_stream, format="png")
            plt.close()
            img_stream.seek(0)
            img = Image(img_stream)
            # Determine where to place the image in the sheet
            if skill == "serve":
                ws.add_image(img, "E2")  # Adjust the cell position as needed
            else:
                ws.add_image(img, "E38")  # Place the second image below the first

        # Save the workbook
        self.workbook.save()


class WorkbookHandler:
    def __init__(self, output_path: str):
        self.wb = self.init_workbook()
        self.output_path = output_path

    def init_workbook(self):
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        return wb

    def create_student_sheet(self, sheet: Student) -> Any:
        ws = self.wb.create_sheet(sheet["name"])
        ws.append(["Name", sheet["name"]])
        ws.append(["Line ID", sheet["line_id"]])
        ws.append(["Handedness", sheet["handedness"]])
        ws.append([])  # Empty row
        ws.append(["Date", "Skill", "Score", "AI Note", "Preview Note", "Reflection"])
        return ws

    def add_student_records(self, ws: Any, portfolio_df: pd.DataFrame):
        for row in portfolio_df.itertuples(index=False):
            ws.append(row)

    def mark_missing_fields(self, ws: Any):
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        # Apply the formatting rule for blanks in columns A to E, starting from row 5
        ws.conditional_formatting.add(
            f"A5:E{ws.max_row}",
            CellIsRule(operator="equal", formula=['""'], fill=red_fill),
        )

    def save(self):
        self.wb.save(self.output_path)
